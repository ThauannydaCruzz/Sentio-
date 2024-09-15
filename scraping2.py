#Importa as bibliotecas necessárias
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import re
import datetime
import pandas as pd
import numpy as np
import openpyxl

#Define o caminho para o chromedriver
caminho_chromedriver = 'C:/Users/Usuario/Documents/chromedriver-win64/chromedriver.exe'

# Configurações para o ChromeDriver
opcoes = Options()
# opcoes.add_argument('--headless')  # Descomente se quiser rodar o Chrome em modo headless (sem interface gráfica)

# Instancia o serviço do ChromeDriver
servico = Service(caminho_chromedriver)

# Instancia o Google Chrome
navegador = webdriver.Chrome(service=servico, options=opcoes)

#Seleciona a página desejada
navegador.get("https://www.reclameaqui.com.br/empresa/{}/".format('itau'))

#Cria um intervalo de espera de 2 segundos
time.sleep(2)

#Localiza o botão para aceitar cookies e clica
navegador.find_element("xpath",'/html/body/div[2]/div[2]/a[1]').click()

#Rola a página 300 pixels para baixo, para que os alvos de cliques futuros não fiquem obstruídos
navegador.execute_script("window.scrollBy(0, 1000)")

#Define os períodos de dados desejados e seus caminhos html
periodos = {'seis_meses':'//*[@id="newPerformanceCard-tab-1"]',
           'doze_meses':'//*[@id="newPerformanceCard-tab-2"]',
           'geral':'//*[@id="newPerformanceCard-tab-5"]'
          }

#Define os elementos a serem coletados e seus caminhos html
elementos = {'nota_geral':'//*[@id="ra-new-reputation"]/span/b',
            'num_reclamacoes':'//*[@id="newPerformanceCard"]/div[2]/div[1]/span',
            'num_nao_respondidas':'//*[@id="newPerformanceCard"]/div[2]/div[3]/span',  
            'perc_recl_resp':'//*[@id="newPerformanceCard"]/div[2]/div[2]/span',
            'novam_negoc':'//*[@id="newPerformanceCard"]/div[2]/div[5]/span', 
            'indice_solucao':'//*[@id="newPerformanceCard"]/div[2]/div[6]/span',  
            'nota_consumidor':'//*[@id="newPerformanceCard"]/div[2]/div[4]/span/strong[2]'
           }

#Cria listas que conterão os dados de cada indicador
listas = {'nota_geral':[],            
          'num_reclamacoes':[],            
          'num_nao_respondidas':[],              
          'perc_recl_resp':[],            
          'novam_negoc':[],             
          'indice_solucao':[],              
          'nota_consumidor':[]
         }

#Define o momento da coleta dos dados
agora = datetime.datetime.now()

#Percorre os períodos definidos
for periodo in periodos:
    #Cria um intervalo de espera de 1 segundo
    time.sleep(1)

    
    # Defina o tempo de espera
    tempo_espera = 10

    # Espere até que o elemento esteja presente e clicável
    elemento = WebDriverWait(navegador, tempo_espera).until(
        EC.element_to_be_clickable((By.XPATH, periodos[periodo]))
     )
    
    #Seleciona o período desejado localizando-o pelo 'xpath' e clicando
    navegador.find_element("xpath",periodos[periodo]).click()
    
    #Cria um intervalo de espera de 1 segundo
    time.sleep(1)
    
    #Percorre todos os elementos
    for elemento in elementos:
        #Localiza um elemento na página pelo 'xpath'
        element = navegador.find_element("xpath",elementos[elemento])

        #Extrai o texto do elemento localizado
        text = element.text

        #Adiciona o texto resultante à lista respectiva
        listas[elemento].append(text)
        
#Fecha o navegador
navegador.quit()

#Cria o dataframe que conterá os dados coletados de forma estruturada
df_resumo = pd.DataFrame(listas)
#Cria uma coluna com a data da coleta dos dados
df_resumo['data'] = agora.date()
#Cria uma coluna com a hora da coleta dos dados
df_resumo['hora'] = agora.time()
#Cria uma coluna que define a qual período pertence cada linha de dados do dataframe
periodos = ['Últimos 6 meses','Últimos 12 meses','Geral']
df_resumo['periodo'] = periodos
#Organiza as colunas do dataframe resultante
df_resumo = df_resumo.iloc[:,[7,8,9,0,1,2,3,4,5,6]]

#Instancia o arquivo xlsx que contém os registros históricos dos indicadores coletados (deve seguir a mesma estrutura do dataframe)
workbook = openpyxl.load_workbook('C:/Users/Usuario/Documents/Projeto_Reclame_Aqui/Reclamacoes.xlsx')

sheet = workbook.create_sheet('Reclamacoes')  # Cria uma nova planilha chamada "Reclamacoes"
workbook.save('C:/Users/Usuario/Documents/ReclameAqui.xlsx')  # Salva o arquivo com a nova planilha

# Instancia o arquivo xlsx que contém os registros históricos dos indicadores coletados
caminho_arquivo = 'C:/Users/Usuario/Documents/Projeto_Reclame_Aqui/Reclamacoes.xlsx'

# Carrega o workbook existente ou cria um novo se não existir
try:
    workbook = openpyxl.load_workbook(caminho_arquivo)
    print("Arquivo carregado com sucesso.")
except FileNotFoundError:
    workbook = openpyxl.Workbook()  # Cria um novo workbook se o arquivo não existir
    print("Arquivo não encontrado. Um novo arquivo foi criado.")

# Verifica se a planilha "Reclamacoes" já existe, se não, cria uma nova
if 'Reclamacoes' not in workbook.sheetnames:
    sheet = workbook.create_sheet('Reclamacoes')  # Cria uma nova planilha chamada "Reclamacoes"
else:
    sheet = workbook['Reclamacoes']  # Seleciona a planilha existente

# Função para verificar e converter valores percentuais
def convert_to_percentage(num):
    try:
        # Expressão regular para capturar apenas o número seguido por '%'
        match = re.search(r'(\d+(\.\d+)?)%', str(num))
        if match:
            return float(match.group(1)) / 100  # Converte o valor capturado em decimal
        else:
            return None  # Retorna None se não houver o símbolo de percentual
    except Exception as e:
        print(f"Erro ao converter '{num}': {e}")
        return None

# Função para limpar strings antes de converter para float
def clean_value(value):
    match = re.search(r'\d+(\.\d+)?', str(value))  # Captura o número, se houver
    if match:
        return float(match.group(0))
    else:
        return None  # Retorna None se não encontrar um número válido

# Converte a coluna 'data' para o formato de data
df_resumo['data'] = df_resumo['data'].apply(lambda data: pd.to_datetime(data).date())

# Percorre as colunas que devem ser formatadas como float e aplica a função de limpeza
for i in range(3, 6):
    df_resumo.iloc[:, i] = df_resumo.iloc[:, i].apply(clean_value)

# Percorre as colunas que devem ser formatadas como percentuais e aplica a função de conversão
for i in range(6, 9):
    df_resumo.iloc[:, i] = df_resumo.iloc[:, i].apply(convert_to_percentage)

# Converte a última coluna em float, aplicando a função de limpeza
df_resumo.iloc[:, 9] = df_resumo.iloc[:, 9].apply(clean_value)

# Insere cada linha do DataFrame no arquivo de registros
for linha in df_resumo.values.tolist():
    sheet.append(linha)

# Salva as alterações no arquivo Excel no local correto
workbook.save(caminho_arquivo)
print("Dados salvos com sucesso.")