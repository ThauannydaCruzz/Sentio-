# Importa as bibliotecas necessárias
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import datetime
import pandas as pd
import openpyxl
import random
from openpyxl.styles import Font, Border, Side


caminho_chromedriver = 'C:/Users/Usuario/Documents/chromedriver-win64/chromedriver.exe'


opcoes = Options()


servico = Service(caminho_chromedriver)


navegador = webdriver.Chrome()

# Lista de bancos que deseja monitorar
bancos = ['santander']


periodos = {
    'seis_meses': '//*[@id="newPerformanceCard-tab-1"]',
    'doze_meses': '//*[@id="newPerformanceCard-tab-2"]',
    '2023': '//*[@id="newPerformanceCard-tab-3"]',
    '2022': '//*[@id="newPerformanceCard-tab-4"]',
    'geral': '//*[@id="newPerformanceCard-tab-5"]'
}


elementos = {
    'nota_geral': '//*[@id="ra-new-reputation"]/span/b',
    'num_reclamacoes': '//*[@id="newPerformanceCard"]/div[2]/div[1]/span',
    'num_nao_respondidas': '//*[@id="newPerformanceCard"]/div[2]/div[3]/span',
    'perc_recl_resp': '//*[@id="newPerformanceCard"]/div[2]/div[2]/span',
    'novam_negoc': '//*[@id="newPerformanceCard"]/div[2]/div[5]/span',
    'indice_solucao': '//*[@id="newPerformanceCard"]/div[2]/div[6]/span',
    'nota_consumidor': '//*[@id="newPerformanceCard"]/div[2]/div[4]/span/strong[2]'
}


def tentar_clicar(xpath, tentativas = 3):
    for i in range(tentativas):
        try:
            WebDriverWait(navegador, 10).until(EC.element_to_be_clickable((By.XPATH, xpath))).click()
            return True
        except Exception as e:
            print(f"Tentativa {i + 1} falhou: {e}")
            time.sleep(random.uniform(2, 5))
    return False


agora = datetime.datetime.now()

# Caminho do arquivo Excel
caminho_arquivo = 'C:/Users/Usuario/Documents/Projeto_Reclame_Aqui/Reclamacoes.xlsx'


for banco in bancos:
    listas = {chave: [] for chave in elementos}  
    listas['banco'] = []  

    print(f"Coletando dados para o banco: {banco}")

   
    url = f"https://www.reclameaqui.com.br/empresa/{banco}/"
    navegador.get(url)
    time.sleep(random.uniform(2, 4))  

    
    if tentar_clicar('/html/body/div[2]/div[2]/a[1]'):
        print("Cookies aceitos.")
    else:
        print("Erro ao aceitar cookies.")

    
    navegador.execute_script("window.scrollBy(0, 700)")
    time.sleep(random.uniform(1, 3))  

    # Percorre os períodos e coleta os dados
    for periodo, xpath_periodo in periodos.items():
        if tentar_clicar(xpath_periodo):
            for nome, xpath in elementos.items():
                try:
                    element = WebDriverWait(navegador, 10).until(EC.presence_of_element_located((By.XPATH, xpath)))
                    listas[nome].append(element.text)
                except Exception as e:
                    listas[nome].append("N/A")
                    print(f"Erro ao coletar {nome} para {banco}: {e}")
        else:
            print(f"Erro ao acessar dados do período {periodo} para o banco {banco}.")

        
        listas['banco'].append(banco)

    # Cria o DataFrame com os dados coletados
    df_resumo = pd.DataFrame(listas)
    df_resumo['data'] = agora.date()
    df_resumo['hora'] = agora.time()
    df_resumo['periodo'] = ['Últimos 6 meses', 'Últimos 12 meses', '2023', '2022', 'Geral']

    
    df_resumo = df_resumo.iloc[:, [7, 8, 9, 10, 0, 1, 2, 3, 4, 5, 6]]

    
    try:
        workbook = openpyxl.load_workbook(caminho_arquivo)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Verifica se a planilha "Reclamacoes" já existe, caso contrário, cria
    if 'Reclamacoes' not in workbook.sheetnames:
        sheet = workbook.create_sheet('Reclamacoes')
    else:
        sheet = workbook['Reclamacoes']


    for linha in df_resumo.values.tolist():
        sheet.append(linha)

    
    for cell in sheet[1]:
        cell.font = Font(bold=True)  

    
    column_widths = {
        'A': 80, 'B': 80, 'C': 80, 'D': 80, 'E': 80,
        'F': 80, 'G': 80, 'H': 80, 'I': 80, 'J': 80, 'K': 80
    }
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = border

    
    workbook.save(caminho_arquivo)
    print(f"Dados do banco {banco} salvos com sucesso.")


navegador.quit()